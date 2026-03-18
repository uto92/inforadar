import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE

wb = openpyxl.Workbook()

# ─── カラーパレット ───────────────────────────────────────────
HDR_DARK   = "1F3864"   # 濃紺（見出し背景）
HDR_MID    = "2E75B6"   # 中青（副見出し）
HDR_LIGHT  = "D6E4F0"   # 薄青（カラム見出し）
ROW_EVEN   = "EBF3FA"   # 偶数行
ROW_ODD    = "FFFFFF"   # 奇数行

# 会社ごとのアクセントカラー
CO_COLORS = {
    "小田急": "E60012",   # 赤
    "東武":   "F5A623",   # 橙
    "東急":   "D50032",   # 臙脂
    "西武":   "003087",   # 紺
    "京王":   "9B1744",   # えんじ
    "JR西日本":"006DB8",  # JRブルー
    "JR東日本":"00B140",  # JRグリーン
}

WHITE  = "FFFFFF"
BLACK  = "000000"
GRAY   = "F2F2F2"
DGRAY  = "595959"

def _bd(style="thin", color="BFBFBF"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_cell(ws, row, col, value, bg=HDR_DARK, fg=WHITE, bold=True, sz=10,
             wrap=True, halign="center", valign="center", colspan=1, rowspan=1):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(color=fg, bold=bold, size=sz, name="游ゴシック")
    cell.alignment = Alignment(horizontal=halign, vertical=valign,
                                wrap_text=wrap)
    cell.border = _bd("thin", "FFFFFF")
    if colspan > 1 or rowspan > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row+rowspan-1, end_column=col+colspan-1)
    return cell

def data_cell(ws, row, col, value, bg=WHITE, fg=BLACK, bold=False, sz=9,
              halign="center", valign="center", wrap=True, italic=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(color=fg, bold=bold, size=sz, name="游ゴシック",
                     italic=italic)
    cell.alignment = Alignment(horizontal=halign, vertical=valign,
                                wrap_text=wrap)
    cell.border = _bd("thin", "D9D9D9")
    return cell

def note_cell(ws, row, col, value, colspan=1):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = PatternFill("solid", fgColor=GRAY)
    cell.font = Font(color=DGRAY, bold=False, size=7, name="游ゴシック",
                     italic=True)
    cell.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
    cell.border = _bd("thin", "CFCFCF")
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+colspan-1)
    return cell


# ═══════════════════════════════════════════════════════════════
# SHEET 1 ── 体制比較
# ═══════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "①体制比較"
ws1.sheet_view.showGridLines = False
ws1.freeze_panes = "C4"

COMPANIES = ["小田急", "東武", "東急", "西武", "京王", "JR西日本", "JR東日本"]
N = len(COMPANIES)

# 列幅
ws1.column_dimensions["A"].width = 4
ws1.column_dimensions["B"].width = 22
for i, co in enumerate(COMPANIES, start=3):
    ws1.column_dimensions[get_column_letter(i)].width = 18

# ── タイトル行 ──────────────────────────────────────
hdr_cell(ws1, 1, 1, "関東私鉄・JR カスタマーサポート体制 比較表（2025年3月版）",
         bg=HDR_DARK, sz=13, bold=True, colspan=N+2)

# ── 会社見出し ──────────────────────────────────────
hdr_cell(ws1, 2, 1, "No.", bg=HDR_MID, sz=9)
hdr_cell(ws1, 2, 2, "比較項目", bg=HDR_MID, sz=9)
for i, co in enumerate(COMPANIES, start=3):
    hdr_cell(ws1, 2, i, co, bg=CO_COLORS[co], sz=10, bold=True)
for i, co in enumerate(COMPANIES, start=3):
    hdr_cell(ws1, 3, i, "← 各社詳細 →", bg=HDR_LIGHT, fg=HDR_DARK, sz=8, bold=False)
hdr_cell(ws1, 3, 1, "", bg=HDR_LIGHT)
hdr_cell(ws1, 3, 2, "", bg=HDR_LIGHT)

# ── データ定義 ──────────────────────────────────────
#  (カテゴリラベル, [(行ラベル, [co1, co2, ...]), ...])
sections = [
    ("【基本チャネル】", [
        ("電話対応",
         ["廃止済（チャット一本化）", "○ 有人電話", "○ 0570-026-109", "○ 有人電話", "○ 03-3325-6644", "○ 有人電話（多拠点）", "○ えきねっとサポートセンター"]),
        ("メール／フォーム",
         ["○ AI FAQフォーム自動化", "○", "○ Q&Aサイト経由", "○", "○", "○", "○ ご意見・ご要望フォーム"]),
        ("LINE公式アカウント",
         ["—", "—", "—", "—", "—", "—", "◎ JR東日本Chat Bot（友だち100万人超）"]),
        ("受付時間（電話）",
         ["廃止", "9:00〜18:00（目安）", "9:30〜17:00", "9:00〜17:00（目安）", "9:00〜18:00", "センター時間内", "センター時間内"]),
    ]),
    ("【AIチャット・ボット】", [
        ("AIチャットボット",
         ["◎ ユーザーローカル製（2023/4〜）", "○ ユーザーローカル製", "○ Salesforce基盤", "不明", "不明", "◎ MOBI BOT（モビルス、2025/2〜）", "○ OKBIZ for Chat&Bot（えきねっと、2019〜）"]),
        ("自動応答チャネル",
         ["Web（シナリオ＋フリーテキスト）", "Web FAQ", "Salesforce Digital Engagement", "—", "—", "電話・Web・メール 3ch統合", "Web＋LINE Bot＋電話AI（2025/7〜）"]),
        ("24h対応",
         ["◎ 24h/365日", "○ FAQ 24h", "× 業時間内のみ", "不明", "×", "○ チャット 24h（一部）", "○ チャット・LINEBot 24h"]),
        ("駅AI端末",
         ["—", "◎ AIさくらさん（6駅、2025/4〜）", "—", "—", "—", "—", "○ どこトレダイヤル（電話AI、2025/7〜）"]),
    ]),
    ("【生成AI・高度DX】", [
        ("生成AI活用",
         ["—", "○ AIさくらさん（生成AIベース）", "—", "—", "—", "◎ ELYZA×ChatGPT（2023/3〜）", "○ 鉄道版生成AI開発中"]),
        ("VoC分析AI",
         ["—", "—", "—", "—", "—", "◎ ELYZA VoC分析パッケージ（2024/12〜）", "△ JRE AI Chat（社内向け）"]),
        ("FAQ自動生成",
         ["—", "—", "—", "—", "—", "◎ 2025年度導入予定", "—"]),
        ("画面共有サポート",
         ["—", "—", "◎ Withdesk Browse（コブラウズ）", "—", "—", "—", "—"]),
        ("多言語対応",
         ["—", "◎ 日英中韓 4言語", "—", "—", "—", "◎ 日英中韓 4言語（有人チャット）", "—"]),
    ]),
    ("【組織・規模】", [
        ("オペレーター数",
         ["—", "—", "約50名（コールセンター）", "—", "—", "544名（兵庫・広島 2拠点）", "—"]),
        ("1日の問い合わせ件数",
         ["—", "—", "TOKYU POINT関連 月約6,000件", "—", "—", "約6,000件（3ch合計）/ 先行拠点で約2,500件", "えきねっとで年間数十万件規模（推計）"]),
        ("月間電話件数",
         ["電話廃止", "—", "約6,000件（TOKYU POINT関連）", "—", "—", "約7万件", "—"]),
        ("AIベンダー",
         ["ユーザーローカル", "ユーザーローカル＋ティファナ", "Salesforce／Withdesk", "—", "—", "ELYZA／モビルス", "オウケイウェイヴ（OKWAVE）"]),
    ]),
    ("【DX成熟度評価】", [
        ("総合DXスコア（★5段階）",
         ["★★★★☆", "★★★★☆", "★★★☆☆", "★★☆☆☆", "★☆☆☆☆", "★★★★★", "★★★★☆"]),
        ("チャネル一本化",
         ["◎ 電話廃止・完全チャット化", "△ 電話残存", "△ 電話残存", "不明", "× 電話中心", "○ 3ch統合管理", "○ 複数ch整備済"]),
        ("リアルタイムデータ活用",
         ["○ 学習型FAQ改善", "○ 駅AIリアルタイム案内", "△", "—", "—", "◎ VoC全件分析（AI）", "○ 運行情報自動応答"]),
    ]),
]

row = 4
for sec_label, items in sections:
    # セクション見出し
    cell = ws1.cell(row=row, column=1, value="")
    cell.fill = PatternFill("solid", fgColor=HDR_MID)
    cell.border = _bd("thin", "FFFFFF")
    c2 = ws1.cell(row=row, column=2, value=sec_label)
    c2.fill = PatternFill("solid", fgColor=HDR_MID)
    c2.font = Font(color=WHITE, bold=True, size=9, name="游ゴシック")
    c2.alignment = Alignment(horizontal="left", vertical="center")
    c2.border = _bd("thin", "FFFFFF")
    for ci in range(3, N+3):
        ws1.cell(row=row, column=ci).fill = PatternFill("solid", fgColor=HDR_MID)
        ws1.cell(row=row, column=ci).border = _bd("thin", "FFFFFF")
    ws1.row_dimensions[row].height = 18
    row += 1

    for idx, (label, vals) in enumerate(items):
        bg = ROW_EVEN if idx % 2 == 0 else ROW_ODD
        data_cell(ws1, row, 1, idx+1, bg=bg, sz=8)
        data_cell(ws1, row, 2, label, bg=bg, halign="left", bold=True, sz=9)
        for ci, val in enumerate(vals, start=3):
            co = COMPANIES[ci-3]
            # ◎は会社カラーで強調
            if "◎" in str(val):
                data_cell(ws1, row, ci, val, bg=bg, fg=CO_COLORS[co], bold=True, sz=8)
            elif "×" in str(val) or "不明" in str(val):
                data_cell(ws1, row, ci, val, bg=bg, fg="808080", sz=8)
            elif "廃止" in str(val):
                data_cell(ws1, row, ci, val, bg=bg, fg="C00000", bold=True, sz=8)
            else:
                data_cell(ws1, row, ci, val, bg=bg, sz=8)
        ws1.row_dimensions[row].height = 40
        row += 1

# ソース注記
row += 1
note_cell(ws1, row, 1,
    "【主要出典】 小田急×ユーザーローカル: userlocal.jp/press/20240227oc/ ／ "
    "東武×AIさくらさん: tifana.ai/news/20250317 ／ "
    "東急×Withdesk: withdesk.com/cases/6065f21e ／ "
    "JR西日本×ELYZA: itmedia.co.jp/business/articles/2412/04/news157.html ／ "
    "JR東日本 どこトレダイヤル: jreast.co.jp/press/2025/20250715_ho01.pdf ／ "
    "えきねっとAIBot: callcenter-japan.com/news_topics/3963.html",
    colspan=N+2)
ws1.row_dimensions[row].height = 30


# ═══════════════════════════════════════════════════════════════
# SHEET 2 ── KPI比較
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("②KPI比較")
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = "C4"

ws2.column_dimensions["A"].width = 4
ws2.column_dimensions["B"].width = 26
for i in range(3, N+3):
    ws2.column_dimensions[get_column_letter(i)].width = 18

hdr_cell(ws2, 1, 1,
    "カスタマーサポート KPI比較表（2025年3月版）",
    bg=HDR_DARK, sz=13, bold=True, colspan=N+2)

hdr_cell(ws2, 2, 1, "No.", bg=HDR_MID, sz=9)
hdr_cell(ws2, 2, 2, "KPI指標", bg=HDR_MID, sz=9)
for i, co in enumerate(COMPANIES, start=3):
    hdr_cell(ws2, 2, i, co, bg=CO_COLORS[co], sz=10, bold=True)
hdr_cell(ws2, 3, 1, "", bg=HDR_LIGHT)
hdr_cell(ws2, 3, 2, "（定義・目標水準）", bg=HDR_LIGHT, fg=HDR_DARK, sz=8, halign="left", bold=False)
for i in range(3, N+3):
    hdr_cell(ws2, 3, i, "公開値 / 推計値", bg=HDR_LIGHT, fg=HDR_DARK, sz=8, bold=False)

kpi_sections = [
    ("【応答品質 KPI】", [
        ("FCR 一次解決率\n（First Call Resolution）\n目安: 70〜90%、優良85%超",
         ["非公開\n（AI FAQで大幅改善中）",
          "非公開\n（AI案内で向上中）",
          "非公開\n（Withdesk活用で改善）",
          "非公開",
          "非公開",
          "★ 54%業務時間削減（介助申請）\n→ 実質FCR向上",
          "非公開\n（AIBot導入で自動解決増）"]),
        ("応答率\n（Response Rate）\n目安: 90%以上、優良95%以上",
         ["廃止のため電話応答率なし\n（チャット24h対応に移行）",
          "非公開",
          "非公開",
          "非公開",
          "非公開",
          "非公開\n（月7万件体制維持）",
          "非公開"]),
        ("チャットボット自己解決率\n（Deflection Rate）\n目安: 30〜50%以上",
         ["◎ 個別回答数 対前年比大幅削減\n→ 推計40〜50%台",
          "○ FAQ誘導で削減中",
          "○ Withdesk活用でWeb誘導強化",
          "—",
          "—",
          "◎ 全問い合わせの一定割合をAI処理\n（VoC分析は全件化達成）",
          "○ えきねっとBot稼働中"]),
        ("VoC（顧客の声）分析カバー率\n目安: 100%が理想",
         ["—",
          "—",
          "—",
          "—",
          "—",
          "◎ AI導入前 約3%→ 導入後 全件分析",
          "—"]),
    ]),
    ("【効率 KPI】", [
        ("AHT 平均処理時間\n（Average Handling Time）\n目安: 4〜7分（音声）",
         ["電話廃止のため非該当\n（チャット対応中心）",
          "非公開",
          "非公開",
          "非公開",
          "非公開",
          "★ 1件あたり65秒削減（要約業務）\n→ AHT大幅短縮",
          "非公開"]),
        ("ACW 平均後処理時間\n（After Call Work）\n目安: 30〜60秒",
         ["—", "—", "—", "—", "—",
          "★ AI要約で大幅削減\n（週報: 2h→30分）",
          "—"]),
        ("ASA 平均応答速度\n（Average Speed of Answer）\n目安: 20秒以内",
         ["廃止（電話なし）",
          "非公開",
          "非公開",
          "非公開",
          "非公開",
          "非公開",
          "非公開"]),
        ("CPH 時間当たり処理件数\n（Calls Per Hour）\n目安: 8〜12件/h",
         ["—（チャット中心）",
          "非公開",
          "非公開",
          "非公開",
          "非公開",
          "非公開\n（AI支援でCPH向上推計）",
          "非公開"]),
        ("メール対応工数削減率",
         ["◎ 大幅削減（フォーム自動化）",
          "—",
          "—",
          "—",
          "—",
          "◎ 業務時間削減 最大54%（駅案内18%・介助54%）",
          "—"]),
    ]),
    ("【顧客満足 KPI】", [
        ("CSAT 顧客満足度スコア\n（Customer Satisfaction）\n目安: 80%以上、優良85%超",
         ["非公開\n（チャット満足度向上中）",
          "○ AIさくら導入で向上中",
          "○ Withdesk導入後「ありがとう」増加",
          "非公開",
          "非公開",
          "非公開\n（VoC品質改善サイクル構築済）",
          "非公開"]),
        ("NPS 顧客推奨度\n（Net Promoter Score）\n目安: 業種平均 +20〜+40",
         ["非公開", "非公開", "非公開", "非公開", "非公開", "非公開", "非公開"]),
        ("有人チャット CSAT\n目安: 85%以上",
         ["—",
          "—",
          "○ コブラウズ導入でCS向上確認",
          "—",
          "—",
          "◎ 有人チャット導入後 忘れ物問合せ ▲24%効率化",
          "—"]),
    ]),
    ("【コスト KPI】", [
        ("CPC コスト・パー・コール\n（Cost Per Call）\n日本平均: 150〜1,000円/件",
         ["廃止（電話コスト=0）\n→チャット中心に移行",
          "非公開",
          "非公開",
          "非公開",
          "非公開",
          "◎ AI活用で人件費効率化推進中\n（544名体制、月7万件）",
          "非公開"]),
        ("チャネル別コスト比較（参考値）\n※グローバル標準 Gartner/ICMI",
         ["電話: ¥1,000〜2,000/件\nチャット: ¥550〜600/件\nAIセルフ: ¥280/件\n→ 小田急は電話廃止でコスト最適化",
          "電話+AI駅端末で分散",
          "電話+有人チャット",
          "—",
          "電話中心（コスト高）",
          "電話: 月7万件×推計¥1,000=\n推計月額7,000万円規模\n→ AI活用で削減推進",
          "電話+チャット+LINE Bot"]),
        ("オペレーター1名月額コスト\n（人件費目安）",
         ["削減済み（チャット移行）",
          "非公開",
          "約20〜30万円×50名\n= 月約1,000〜1,500万円",
          "非公開",
          "非公開",
          "約20〜30万円×544名\n= 月推計1.1〜1.6億円",
          "非公開"]),
        ("AI投資対効果（ROI目安）\n※業界平均: 投資回収12〜18ヶ月",
         ["◎ チャット一本化で\n電話インフラコスト撤廃",
          "○ 駅AI実証中（ROI検証段階）",
          "○ Withdesk導入後CSコスト低減",
          "—",
          "—",
          "◎ 要約65秒×月7万件\n= 月約7,500時間削減（推計）",
          "○ LINEBot 100万友だちで\n問合せ分散効果"]),
    ]),
    ("【業界標準ベンチマーク（参考）】", [
        ("FCR 業界標準",
         ["一般: 70〜85%"]*N),
        ("優良水準 FCR",
         ["AI活用優良: 85〜90%超"]*N),
        ("AHT 業界標準（音声）",
         ["4〜7分"]*N),
        ("CSAT 業界標準",
         ["80%以上 / 優良85%超"]*N),
        ("ASA 業界標準",
         ["20秒以内"]*N),
        ("電話 CPC（グローバル）",
         ["$6.69〜$13.50（¥1,000〜2,000）"]*N),
        ("有人チャット CPC（グローバル）",
         ["$3.64〜$4.00（¥550〜600）"]*N),
        ("AIセルフ CPC（Gartner中央値）",
         ["$1.84（¥280）"]*N),
        ("FAQ解決 CPC（最安）",
         ["$0.10以下（¥15以下）"]*N),
    ]),
]

row = 4
for sec_label, items in kpi_sections:
    cell = ws2.cell(row=row, column=1, value="")
    cell.fill = PatternFill("solid", fgColor=HDR_MID)
    cell.border = _bd("thin", "FFFFFF")
    c2 = ws2.cell(row=row, column=2, value=sec_label)
    c2.fill = PatternFill("solid", fgColor=HDR_MID)
    c2.font = Font(color=WHITE, bold=True, size=9, name="游ゴシック")
    c2.alignment = Alignment(horizontal="left", vertical="center")
    c2.border = _bd("thin", "FFFFFF")
    for ci in range(3, N+3):
        ws2.cell(row=row, column=ci).fill = PatternFill("solid", fgColor=HDR_MID)
        ws2.cell(row=row, column=ci).border = _bd("thin", "FFFFFF")
    ws2.row_dimensions[row].height = 18
    row += 1

    for idx, (label, vals) in enumerate(items):
        bg = ROW_EVEN if idx % 2 == 0 else ROW_ODD
        data_cell(ws2, row, 1, idx+1, bg=bg, sz=8)
        data_cell(ws2, row, 2, label, bg=bg, halign="left", bold=True, sz=8)
        for ci, val in enumerate(vals, start=3):
            co = COMPANIES[ci-3]
            if "★" in str(val) or "◎" in str(val):
                data_cell(ws2, row, ci, val, bg=bg, fg=CO_COLORS[co], bold=True, sz=8)
            elif "非公開" in str(val) or "—" == str(val).strip():
                data_cell(ws2, row, ci, val, bg=bg, fg="909090", sz=8, italic=True)
            elif "廃止" in str(val):
                data_cell(ws2, row, ci, val, bg=bg, fg="C00000", bold=True, sz=8)
            else:
                data_cell(ws2, row, ci, val, bg=bg, sz=8)
        ws2.row_dimensions[row].height = 52
        row += 1

# ソース
row += 1
note_cell(ws2, row, 1,
    "【KPIデータ出典】 "
    "JR西日本×ELYZA KPI: itmedia.co.jp/business/articles/2412/04/news157.html ／ "
    "小田急メール削減: userlocal.jp/press/20240227oc/ ／ "
    "東急Withdesk: withdesk.com/cases/6065f21e ／ "
    "チャネル別コスト(Gartner): icmi.com/resources/2021/contact-center-metric-cost-per-contact ／ "
    "業界ベンチマーク: lorikeetcx.ai/articles/contact-center-benchmarks ／ "
    "コールセンターKPI: mobilus.co.jp/lab/trend/contactcenter-kpi-1/ ／ "
    "CPC解説: kddimessagecast.jp/blog/ai_ivr_automation/240808/ ／ "
    "日本コスト相場: tactinc.jp/article/callceter-cpc",
    colspan=N+2)
ws2.row_dimensions[row].height = 40


# ═══════════════════════════════════════════════════════════════
# SHEET 3 ── コスト試算
# ═══════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("③コスト試算")
ws3.sheet_view.showGridLines = False

ws3.column_dimensions["A"].width = 32
ws3.column_dimensions["B"].width = 20
ws3.column_dimensions["C"].width = 20
ws3.column_dimensions["D"].width = 20
ws3.column_dimensions["E"].width = 20
ws3.column_dimensions["F"].width = 38

hdr_cell(ws3, 1, 1, "チャネル別コスト試算表（月間10万件問い合わせの場合の試算）",
         bg=HDR_DARK, sz=13, bold=True, colspan=6)

headers = ["チャネル", "1件あたりコスト\n（グローバル参考）", "1件あたりコスト\n（日本目安）",
           "月10万件の場合\n月間総コスト（推計）", "対電話比コスト比",
           "備考・出典"]
for ci, h in enumerate(headers, start=1):
    hdr_cell(ws3, 2, ci, h, bg=HDR_MID, sz=9)

cost_data = [
    ["電話（有人オペレーター）",   "$6.69〜$13.50",   "¥1,000〜¥2,000/件",   "¥1億〜¥2億/月",   "100%（基準）",
     "人件費・通信費・施設費含む。日本の自社運営コール\nセンター（50席規模）では月1,000〜1,500万円/拠点。\n出典: Gartner / tactinc.jp/article/callceter-cpc"],
    ["有人チャット\n（オペレーター）",    "$3.64〜$4.00",    "¥550〜¥600/件",    "¥5,500〜¥6,000万/月",    "約40〜55%",
     "電話の約1/2。同時複数対応が可能なため\n実質コストは更に低下。\n出典: Gartner / mobilus.co.jp/lab/chat-support/billing-cost-reduction-by-chat/"],
    ["AIチャットボット\n（セルフサービス）",  "$0.25〜$0.50",   "¥40〜¥75/件",   "¥400〜¥750万/月",   "約3〜7%",
     "初期導入費用は別途100〜500万円程度。\nユーザーローカル等SaaS型なら月額数十万円〜。\n出典: Gartner中央値$1.84 / userlocal.jp"],
    ["FAQ（セルフ解決）",   "$0.10以下",   "¥15以下/件",   "¥150万以下/月",   "約1%以下",
     "コンテンツ整備・更新人件費が主コスト。\n問い合わせ件数が多いほどROIが高い。\n出典: Gartner / Lorikeet CX Benchmarks 2026"],
    ["AIエージェント\n（生成AI自動対応）",  "$0.50〜$2.00",   "¥75〜¥300/件",   "¥750〜¥3,000万/月",   "約5〜20%",
     "LLM API利用料・システム費含む。\nJR西日本×ELYZA事例: 月7万件×65秒削減\n= 月約7,500時間相当の人件費削減（推計）\n出典: itmedia.co.jp/business/articles/2412/04/news157.html"],
    ["IVR自動音声応答",   "$0.50〜$1.50",   "¥75〜¥225/件",   "¥750〜¥2,250万/月",   "約7〜15%",
     "JR東日本「どこトレダイヤル」（2025/7〜）はこの区分。\nシンプルな情報提供（遅延・走行位置）に特化。\n出典: jreast.co.jp/press/2025/20250715_ho01.pdf"],
]

for ridx, row_data in enumerate(cost_data):
    bg = ROW_EVEN if ridx % 2 == 0 else ROW_ODD
    for ci, val in enumerate(row_data, start=1):
        if ci == 1:
            data_cell(ws3, ridx+3, ci, val, bg=bg, halign="left", bold=True, sz=9)
        elif ci == 5:  # 比率列
            pct = val
            if "100%" in pct:
                data_cell(ws3, ridx+3, ci, val, bg="FFE0E0", fg="C00000", bold=True, sz=9)
            elif "1%" in pct:
                data_cell(ws3, ridx+3, ci, val, bg="E0FFE0", fg="107C10", bold=True, sz=9)
            else:
                data_cell(ws3, ridx+3, ci, val, bg=bg, fg=HDR_MID, bold=True, sz=9)
        elif ci == 6:
            data_cell(ws3, ridx+3, ci, val, bg=GRAY, fg=DGRAY, halign="left", sz=7, italic=True)
        else:
            data_cell(ws3, ridx+3, ci, val, bg=bg, sz=9)
    ws3.row_dimensions[ridx+3].height = 60

# 補足行
r_note = len(cost_data) + 4
note_cell(ws3, r_note, 1,
    "【注意】コスト数値はグローバル標準値（Gartner/ICMI）および日本国内相場を基に算出した参考値です。"
    "実際のコストは業務難易度・オペレーター時給・システム選定・件数規模によって大きく異なります。"
    "日本国内の公式統計はCCAJ（ccaj.or.jp）またはHDI Japan（hdi-japan.com）の有料レポートを参照ください。",
    colspan=6)
ws3.row_dimensions[r_note].height = 35

# JR西日本 実績試算
r2 = r_note + 2
hdr_cell(ws3, r2, 1, "【実績参考】JR西日本カスタマーリレーションズ × ELYZA 生成AI導入効果試算",
         bg=CO_COLORS["JR西日本"], sz=10, bold=True, colspan=6)
jrw_data = [
    ["月間問い合わせ件数（電話+メール）", "約7万件/月", "544名体制（兵庫・広島 2拠点）"],
    ["AI要約による1件あたり削減時間", "約65秒/件", "JR西日本×ELYZA 実証値（2023〜）"],
    ["月間総削減時間（推計）", "約7万件×65秒 ≒ 7,500時間/月", "オペレーター換算で約47名分（160h/月）"],
    ["介助申請業務の処理時間削減率", "54%削減（実証実験）", "2023年5〜8月 実証実験値"],
    ["駅案内業務の処理時間削減率", "18%削減（実証実験）", "2023年5〜8月 実証実験値"],
    ["週報作成時間", "2時間 → 30分（▲75%）", "ELYZA VoC分析パッケージ"],
    ["VoC分析対象割合", "約3% → 全件（▲97pt改善）", "生成AI導入前後比較"],
]
for ridx, row_data in enumerate(jrw_data):
    bg = ROW_EVEN if ridx % 2 == 0 else ROW_ODD
    for ci, val in enumerate(row_data, start=1):
        bold = ci == 1
        data_cell(ws3, r2+ridx+1, ci, val, bg=bg, halign="left", bold=bold, sz=9)
    # 空白セル
    for ci in range(len(row_data)+1, 7):
        data_cell(ws3, r2+ridx+1, ci, "", bg=bg)
    ws3.row_dimensions[r2+ridx+1].height = 22

row_src = r2 + len(jrw_data) + 2
note_cell(ws3, row_src, 1,
    "出典: ITmedia itmedia.co.jp/business/articles/2412/04/news157.html ／ "
    "JR西日本プレス westjr.co.jp/press/article/2024/12/page_26776.html",
    colspan=6)
ws3.row_dimensions[row_src].height = 20


# ─── 保存 ────────────────────────────────────────────
outpath = "/home/user/inforadar/鉄道CS比較表_2025.xlsx"
wb.save(outpath)
print(f"Saved: {outpath}")
